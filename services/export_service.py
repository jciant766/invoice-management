"""
Export Service for generating Excel and PDF files.

Generates Schedule of Payments (Skeda tal-Hlasijiet) in the exact format
required by Malta's Department of Local Government (DLG).
"""

import csv
from io import BytesIO, StringIO
from datetime import datetime
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from services.export_profile_service import normalize_signatory_value


def normalize_signatories(signatories: Optional[dict] = None) -> Dict[str, str]:
    """Normalize signatory dictionary for export rendering."""
    signatories = signatories or {}
    return {
        "sindku": normalize_signatory_value(signatories.get("sindku")) or "",
        "segretarju_ezekuttiv": normalize_signatory_value(signatories.get("segretarju_ezekuttiv")) or "",
        "proponent": normalize_signatory_value(signatories.get("proponent")) or "",
        "sekondant": normalize_signatory_value(signatories.get("sekondant")) or "",
    }


def format_date_value(date_value) -> str:
    """
    Format a date value that may be a string or datetime object.
    Returns DD/MM/YYYY format.
    """
    if not date_value:
        return ""
    if isinstance(date_value, str):
        # Try to parse and reformat
        try:
            if 'T' in date_value:
                # ISO format with time
                dt = datetime.fromisoformat(date_value.replace('Z', '+00:00'))
            else:
                # Just date
                dt = datetime.strptime(date_value, '%Y-%m-%d')
            return dt.strftime('%d/%m/%Y')
        except (ValueError, TypeError):
            return date_value  # Return as-is if parsing fails
    elif hasattr(date_value, 'strftime'):
        return date_value.strftime('%d/%m/%Y')
    return str(date_value)


def sanitize_cell_value(value) -> str:
    """
    Sanitize cell values to prevent Excel formula injection.

    Formula injection occurs when cell content starts with =, @, +, or -
    which Excel interprets as formulas. This can be exploited to execute
    arbitrary formulas or external links.

    Args:
        value: The cell value to sanitize

    Returns:
        Sanitized string safe for Excel
    """
    if value is None:
        return ""

    str_value = str(value)

    # Characters that trigger formula interpretation in Excel
    dangerous_chars = ('=', '@', '+', '-', '\t', '\r', '\n')

    if str_value and str_value[0] in dangerous_chars:
        # Prefix with single quote to force text interpretation
        return "'" + str_value

    return str_value


def generate_schedule_excel(
    invoices: List,
    sitting_number: Optional[str] = None,
    month_year: Optional[str] = None,
    signatories: Optional[dict] = None
) -> BytesIO:
    """
    Generate Excel file matching exact DLG template format.

    Args:
        invoices: List of Invoice objects to export
        sitting_number: Council sitting number (e.g., "23")
        month_year: Month and year string (e.g., "November 2025")

    Returns:
        BytesIO: Excel file as bytes buffer
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Skeda tal-Hlasijiet"
    signatory_values = normalize_signatories(signatories)

    # Define styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")

    # Row 1: Title
    ws.merge_cells('A1:P1')  # 16 columns (A-P)
    ws['A1'] = "KUNSILL LOKALI TAS-SLIEMA"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Row 2: Subtitle
    ws.merge_cells('A2:P2')  # 16 columns (A-P)
    title_text = "SKEDA TAL-HLASIJIET"
    if sitting_number:
        title_text += f" - Seduta Nru. {sitting_number}"
    if month_year:
        title_text += f" - {month_year}"
    ws['A2'] = title_text
    ws['A2'].font = header_font
    ws['A2'].alignment = Alignment(horizontal='center')

    # Row 3: Empty row
    ws.row_dimensions[3].height = 10

    # Row 4: Column headers (Maltese)
    headers = [
        ("A", "Seduta Nru", 10),
        ("B", "Supplier Code", 12),
        ("C", "#", 5),
        ("D", "Fornitur", 30),
        ("E", "Ammont tal-Invoice", 15),
        ("F", "Ammont li ser Jithallas", 18),
        ("G", "Metodu*", 8),
        ("H", "Prokur.", 8),
        ("I", "Deskrizzjoni", 40),
        ("J", "Data tal-Invoice", 15),
        ("K", "Nru. tal-Invoice", 15),
        ("L", "Nru. Tal-PR", 12),
        ("M", "Nru. Tal-PO", 12),
        ("N", "TF/CHQ Nru", 15),
        ("O", "PJV Number", 12),
        ("P", "Status", 10),
    ]

    for col, header_text, width in headers:
        cell = ws[f'{col}4']
        cell.value = header_text
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[col].width = width

    # Row 5+: Invoice data
    row = 5
    total_invoice = 0
    total_payment = 0

    for idx, invoice in enumerate(invoices, start=1):
        # Seduta Nru
        ws[f'A{row}'] = sitting_number or ""
        ws[f'A{row}'].border = border

        # Supplier Code (using supplier ID for now)
        ws[f'B{row}'] = str(invoice.supplier_id)
        ws[f'B{row}'].border = border

        # Sequential number
        ws[f'C{row}'] = idx
        ws[f'C{row}'].border = border

        # Supplier name (with null guard and formula injection protection)
        supplier_name = invoice.supplier.name if invoice.supplier else "Unknown Supplier"
        ws[f'D{row}'] = sanitize_cell_value(supplier_name)
        ws[f'D{row}'].border = border

        # Invoice amount
        ws[f'E{row}'] = float(invoice.invoice_amount)
        ws[f'E{row}'].number_format = '#,##0.00'
        ws[f'E{row}'].border = border
        total_invoice += float(invoice.invoice_amount)

        # Payment amount
        ws[f'F{row}'] = float(invoice.payment_amount)
        ws[f'F{row}'].number_format = '#,##0.00'
        ws[f'F{row}'].border = border
        total_payment += float(invoice.payment_amount)

        # Method of request
        ws[f'G{row}'] = invoice.method_request
        ws[f'G{row}'].border = border
        ws[f'G{row}'].alignment = Alignment(horizontal='center')

        # Procurement method
        ws[f'H{row}'] = invoice.method_procurement
        ws[f'H{row}'].border = border
        ws[f'H{row}'].alignment = Alignment(horizontal='center')

        # Description (sanitized for formula injection)
        ws[f'I{row}'] = sanitize_cell_value(invoice.description)
        ws[f'I{row}'].border = border
        ws[f'I{row}'].alignment = Alignment(wrap_text=True)

        # Invoice date (DD/MM/YYYY format)
        ws[f'J{row}'] = format_date_value(invoice.invoice_date)
        ws[f'J{row}'].border = border
        ws[f'J{row}'].alignment = Alignment(horizontal='center')

        # Invoice number (sanitized)
        ws[f'K{row}'] = sanitize_cell_value(invoice.invoice_number)
        ws[f'K{row}'].border = border

        # PR number (Purchase Request - not tracked in current model)
        ws[f'L{row}'] = ""
        ws[f'L{row}'].border = border

        # PO number (sanitized)
        ws[f'M{row}'] = sanitize_cell_value(invoice.po_number or "")
        ws[f'M{row}'].border = border

        # TF or CHQ number (sanitized)
        tf_chq_value = invoice.chq_number or invoice.tf_number or ""
        ws[f'N{row}'] = sanitize_cell_value(tf_chq_value)
        ws[f'N{row}'].border = border
        ws[f'N{row}'].alignment = Alignment(horizontal='center')

        # PJV number (sanitized)
        ws[f'O{row}'] = sanitize_cell_value(invoice.pjv_number)
        ws[f'O{row}'].border = border

        # Status (Active/Inactive based on is_deleted)
        status_text = "Inactive" if invoice.is_deleted else ("Voided" if getattr(invoice, "is_void", 0) else "Active")
        ws[f'P{row}'] = status_text
        ws[f'P{row}'].border = border
        ws[f'P{row}'].alignment = Alignment(horizontal='center')
        if invoice.is_deleted:
            ws[f'P{row}'].font = Font(color="FF0000")  # Red for inactive

        row += 1

    # Totals row
    ws[f'D{row}'] = "TOTALI:"
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'D{row}'].border = border

    ws[f'E{row}'] = total_invoice
    ws[f'E{row}'].number_format = '#,##0.00'
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'E{row}'].border = border

    ws[f'F{row}'] = total_payment
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'F{row}'].border = border

    # Add borders to empty cells in totals row
    for col in ['A', 'B', 'C', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        ws[f'{col}{row}'].border = border

    # Row height for header
    ws.row_dimensions[4].height = 30

    # Legend row
    row += 2
    ws[f'A{row}'] = "* Metodu: Inv=Invoice, Rec=Receipt, RFP=Request for Payment, PP=Part Payment, DP=Deposit, EC=Expense Claim"
    ws[f'A{row}'].font = Font(italic=True, size=9)

    row += 1
    ws[f'A{row}'] = "  Prokur.: DA=Direct Order Approvata, D=Direct Order, T=Tender, K=Kwotazzjoni, R=Refund"
    ws[f'A{row}'].font = Font(italic=True, size=9)

    # Signature Section
    row += 3

    # Signature boxes header row
    sig_header_fill = PatternFill(start_color="e5e7eb", end_color="e5e7eb", fill_type="solid")
    sig_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    # Create signature section in columns L-P (right side of sheet)
    ws[f'L{row}'] = "IFFIRMATA"
    ws[f'L{row}'].font = Font(bold=True, size=10)
    ws[f'L{row}'].alignment = Alignment(horizontal='center')

    ws[f'N{row}'] = "IFFIRMATA"
    ws[f'N{row}'].font = Font(bold=True, size=10)
    ws[f'N{row}'].alignment = Alignment(horizontal='center')

    row += 1
    # Name boxes
    ws.merge_cells(f'L{row}:M{row}')
    ws[f'L{row}'] = ""  # Empty for handwritten name
    ws[f'L{row}'].border = sig_border
    ws[f'L{row}'].fill = sig_header_fill
    ws.row_dimensions[row].height = 25

    ws.merge_cells(f'N{row}:O{row}')
    ws[f'N{row}'] = ""  # Empty for handwritten name
    ws[f'N{row}'].border = sig_border
    ws[f'N{row}'].fill = sig_header_fill

    row += 1
    # Printed names
    ws.merge_cells(f'L{row}:M{row}')
    ws[f'L{row}'] = signatory_values["sindku"]
    ws[f'L{row}'].alignment = Alignment(horizontal='center')
    ws[f'L{row}'].font = Font(size=9)

    ws.merge_cells(f'N{row}:O{row}')
    ws[f'N{row}'] = signatory_values["segretarju_ezekuttiv"]
    ws[f'N{row}'].alignment = Alignment(horizontal='center')
    ws[f'N{row}'].font = Font(size=9)

    row += 1
    # Role labels
    ws[f'L{row}'] = "Sindku"
    ws[f'L{row}'].font = Font(size=9)
    ws[f'L{row}'].alignment = Alignment(horizontal='center')

    ws[f'N{row}'] = "Segretarju Ezekuttiv"
    ws[f'N{row}'].font = Font(size=9)
    ws[f'N{row}'].alignment = Alignment(horizontal='center')

    row += 2
    # Proposer and Seconder section
    ws[f'L{row}'] = "PROPONENT"
    ws[f'L{row}'].font = Font(bold=True, size=10)
    ws[f'L{row}'].alignment = Alignment(horizontal='center')

    ws[f'N{row}'] = "SEKONDANT"
    ws[f'N{row}'].font = Font(bold=True, size=10)
    ws[f'N{row}'].alignment = Alignment(horizontal='center')

    row += 1
    # Signature boxes for councillors
    ws.merge_cells(f'L{row}:M{row}')
    ws[f'L{row}'] = ""  # Empty for handwritten signature
    ws[f'L{row}'].border = sig_border
    ws[f'L{row}'].fill = sig_header_fill
    ws.row_dimensions[row].height = 35

    ws.merge_cells(f'N{row}:O{row}')
    ws[f'N{row}'] = ""  # Empty for handwritten signature
    ws[f'N{row}'].border = sig_border
    ws[f'N{row}'].fill = sig_header_fill

    row += 1
    # Name labels for councillors
    ws.merge_cells(f'L{row}:M{row}')
    ws[f'L{row}'] = signatory_values["proponent"]
    ws[f'L{row}'].border = sig_border
    ws[f'L{row}'].alignment = Alignment(horizontal='center')
    ws[f'L{row}'].font = Font(size=9)
    ws.row_dimensions[row].height = 20

    ws.merge_cells(f'N{row}:O{row}')
    ws[f'N{row}'] = signatory_values["sekondant"]
    ws[f'N{row}'].border = sig_border
    ws[f'N{row}'].alignment = Alignment(horizontal='center')
    ws[f'N{row}'].font = Font(size=9)

    row += 1
    ws[f'L{row}'] = "(Kunsillier)"
    ws[f'L{row}'].font = Font(size=9)
    ws[f'L{row}'].alignment = Alignment(horizontal='center')

    ws[f'N{row}'] = "(Kunsillier)"
    ws[f'N{row}'].font = Font(size=9)
    ws[f'N{row}'].alignment = Alignment(horizontal='center')

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def generate_schedule_pdf(
    invoices: List,
    sitting_number: Optional[str] = None,
    month_year: Optional[str] = None,
    signatories: Optional[dict] = None
) -> BytesIO:
    """
    Generate PDF file for public viewing.

    Args:
        invoices: List of Invoice objects to export
        sitting_number: Council sitting number
        month_year: Month and year string

    Returns:
        BytesIO: PDF file as bytes buffer
    """
    buffer = BytesIO()
    signatory_values = normalize_signatories(signatories)

    # Create document in landscape
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10*mm,
        leftMargin=10*mm,
        topMargin=15*mm,
        bottomMargin=15*mm
    )

    elements = []
    styles = getSampleStyleSheet()

    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1,  # Center
        spaceAfter=6
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        alignment=1,
        spaceAfter=12
    )

    # Add title
    elements.append(Paragraph("KUNSILL LOKALI TAS-SLIEMA", title_style))

    subtitle_text = "SKEDA TAL-HLASIJIET"
    if sitting_number:
        subtitle_text += f" - Seduta Nru. {sitting_number}"
    if month_year:
        subtitle_text += f" - {month_year}"
    elements.append(Paragraph(subtitle_text, subtitle_style))
    elements.append(Spacer(1, 10*mm))

    # Table headers
    table_headers = [
        "#",
        "Fornitur",
        "Ammont (EUR)",
        "Hlas (EUR)",
        "Met.",
        "Prok.",
        "Deskrizzjoni",
        "Data",
        "Inv. Nru",
        "TF/CHQ",
        "Status"
    ]

    # Table data
    table_data = [table_headers]

    total_invoice = 0
    total_payment = 0

    for idx, invoice in enumerate(invoices, start=1):
        total_invoice += float(invoice.invoice_amount)
        total_payment += float(invoice.payment_amount)

        # Status based on is_deleted flag
        status_text = "Inactive" if invoice.is_deleted else ("Voided" if getattr(invoice, "is_void", 0) else "Active")

        # Supplier name with null guard
        supplier_name = invoice.supplier.name if invoice.supplier else "Unknown Supplier"

        row = [
            str(idx),
            supplier_name[:25] + "..." if len(supplier_name) > 25 else supplier_name,
            f"{float(invoice.invoice_amount):,.2f}",
            f"{float(invoice.payment_amount):,.2f}",
            invoice.method_request,
            invoice.method_procurement,
            invoice.description[:40] + "..." if len(invoice.description) > 40 else invoice.description,
            format_date_value(invoice.invoice_date),
            invoice.invoice_number[:15] if len(invoice.invoice_number) > 15 else invoice.invoice_number,
            invoice.chq_number or invoice.tf_number or "-",
            status_text
        ]
        table_data.append(row)

    # Add totals row
    table_data.append([
        "",
        "TOTALI:",
        f"{total_invoice:,.2f}",
        f"{total_payment:,.2f}",
        "", "", "", "", "", "", ""
    ])

    # Column widths (in mm, converted to points) - adjusted for Status column
    col_widths = [8*mm, 32*mm, 20*mm, 20*mm, 12*mm, 12*mm, 50*mm, 18*mm, 22*mm, 18*mm, 15*mm]

    # Create table
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Table style
    table_style = TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

        # Data rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # # column
        ('ALIGN', (2, 1), (3, -1), 'RIGHT'),   # Amount columns
        ('ALIGN', (4, 1), (5, -1), 'CENTER'),  # Method columns
        ('ALIGN', (7, 1), (7, -1), 'CENTER'),  # Date column
        ('ALIGN', (9, 1), (9, -1), 'CENTER'),  # TF column
        ('ALIGN', (10, 1), (10, -1), 'CENTER'),  # Status column

        # Totals row
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f3f4f6')),

        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9fafb')]),
    ])

    table.setStyle(table_style)
    elements.append(table)

    # Legend
    elements.append(Spacer(1, 10*mm))
    legend_style = ParagraphStyle(
        'Legend',
        parent=styles['Normal'],
        fontSize=7,
        textColor=colors.grey
    )
    elements.append(Paragraph(
        "Met.: Inv=Invoice, Rec=Receipt, RFP=Request for Payment, PP=Part Payment, DP=Deposit, EC=Expense Claim",
        legend_style
    ))
    elements.append(Paragraph(
        "Prok.: DA=Direct Order Approvata, D=Direct Order, T=Tender, K=Kwotazzjoni, R=Refund",
        legend_style
    ))

    # Signature Section
    elements.append(Spacer(1, 15*mm))

    # Signature table with boxes for councillors to sign
    sig_data = [
        ["IFFIRMATA", "", "IFFIRMATA", ""],
        ["", "", "", ""],  # Empty signature boxes
        [signatory_values["sindku"], "", signatory_values["segretarju_ezekuttiv"], ""],
        ["Sindku", "", "Segretarju Ezekuttiv", ""],
        ["", "", "", ""],  # Spacer row
        ["PROPONENT", "", "SEKONDANT", ""],
        ["", "", "", ""],  # Empty signature boxes
        [signatory_values["proponent"], "", signatory_values["sekondant"], ""],
        ["(Kunsillier)", "", "(Kunsillier)", ""],
    ]

    sig_col_widths = [50*mm, 50*mm, 50*mm, 50*mm]
    sig_table = Table(sig_data, colWidths=sig_col_widths)

    sig_style = TableStyle([
        # Header labels
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

        # Signature boxes (row 1 and 5)
        ('BOX', (0, 1), (1, 1), 1, colors.black),
        ('BOX', (2, 1), (3, 1), 1, colors.black),
        ('BOX', (0, 6), (1, 6), 1, colors.black),
        ('BOX', (2, 6), (3, 6), 1, colors.black),
        ('BACKGROUND', (0, 1), (1, 1), colors.HexColor('#f3f4f6')),
        ('BACKGROUND', (2, 1), (3, 1), colors.HexColor('#f3f4f6')),
        ('BACKGROUND', (0, 6), (1, 6), colors.HexColor('#f3f4f6')),
        ('BACKGROUND', (2, 6), (3, 6), colors.HexColor('#f3f4f6')),

        # Printed names
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica'),
        ('FONTSIZE', (0, 2), (-1, 2), 9),
        ('ALIGN', (0, 2), (-1, 2), 'CENTER'),
        ('FONTNAME', (0, 7), (-1, 7), 'Helvetica'),
        ('FONTSIZE', (0, 7), (-1, 7), 9),
        ('ALIGN', (0, 7), (-1, 7), 'CENTER'),

        # Role labels
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica'),
        ('FONTSIZE', (0, 3), (-1, 3), 9),
        ('ALIGN', (0, 3), (-1, 3), 'CENTER'),

        # Proposer/Seconder headers
        ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 5), (-1, 5), 10),
        ('ALIGN', (0, 5), (-1, 5), 'CENTER'),

        # Councillor labels
        ('FONTNAME', (0, 8), (-1, 8), 'Helvetica'),
        ('FONTSIZE', (0, 8), (-1, 8), 9),
        ('ALIGN', (0, 8), (-1, 8), 'CENTER'),

        # Row heights
        ('TOPPADDING', (0, 1), (-1, 1), 15),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 15),
        ('TOPPADDING', (0, 6), (-1, 6), 15),
        ('BOTTOMPADDING', (0, 6), (-1, 6), 15),
    ])

    sig_table.setStyle(sig_style)
    elements.append(sig_table)

    # Footer with generation date
    elements.append(Spacer(1, 5*mm))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=7,
        textColor=colors.grey,
        alignment=2  # Right
    )
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        footer_style
    ))

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    return buffer


def generate_schedule_csv(
    invoices: List,
    sitting_number: Optional[str] = None
) -> BytesIO:
    """
    Generate CSV file for data export.

    Args:
        invoices: List of Invoice objects to export
        sitting_number: Council sitting number

    Returns:
        BytesIO: CSV file as bytes buffer
    """
    output = StringIO()
    writer = csv.writer(output)

    # Headers
    headers = [
        "Sitting No",
        "Supplier Code",
        "#",
        "Supplier",
        "Invoice Amount",
        "Payment Amount",
        "Method Request",
        "Method Procurement",
        "Description",
        "Invoice Date",
        "Invoice Number",
        "PO Number",
        "PJV Number",
        "TF Number",
        "CHQ Number",
        "Status",
        "Approval",
        "Approved Date",
        "Proposer",
        "Seconder",
        "Fiscal Receipt"
    ]
    writer.writerow(headers)

    # Data rows
    for idx, invoice in enumerate(invoices, start=1):
        # Status based on is_deleted flag
        status_text = "Inactive" if invoice.is_deleted else ("Voided" if getattr(invoice, "is_void", 0) else "Active")
        approval_text = "Approved" if invoice.is_approved else "Pending"

        # Supplier name with null guard
        supplier_name = invoice.supplier.name if invoice.supplier else "Unknown Supplier"

        row = [
            sitting_number or "",
            invoice.supplier_id,
            idx,
            sanitize_cell_value(supplier_name),
            float(invoice.invoice_amount),
            float(invoice.payment_amount),
            sanitize_cell_value(invoice.method_request),
            sanitize_cell_value(invoice.method_procurement),
            sanitize_cell_value(invoice.description),
            format_date_value(invoice.invoice_date),
            sanitize_cell_value(invoice.invoice_number),
            sanitize_cell_value(invoice.po_number or ""),
            sanitize_cell_value(invoice.pjv_number),
            sanitize_cell_value(invoice.tf_number or ""),
            sanitize_cell_value(invoice.chq_number or ""),
            status_text,
            approval_text,
            format_date_value(invoice.approved_date),
            sanitize_cell_value(invoice.proposer_councillor or ""),
            sanitize_cell_value(invoice.seconder_councillor or ""),
            "Yes" if invoice.fiscal_receipt_path else "No"
        ]
        writer.writerow(row)

    # Totals row
    total_invoice = sum(float(inv.invoice_amount) for inv in invoices)
    total_payment = sum(float(inv.payment_amount) for inv in invoices)
    writer.writerow([])
    writer.writerow(["", "", "", "TOTAL", total_invoice, total_payment, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])

    # Convert to bytes
    buffer = BytesIO()
    buffer.write(output.getvalue().encode('utf-8-sig'))  # UTF-8 with BOM for Excel compatibility
    buffer.seek(0)

    return buffer


def generate_bulk_vouchers_pdf(invoices: List) -> BytesIO:
    """
    Generate a multi-page PDF with one payment voucher per selected invoice.

    Each invoice is rendered on its own A4 page so users can print in bulk.
    """
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    page_width, page_height = A4

    def wrap_lines(text: str, max_width: float, font_name: str, font_size: float, max_lines: int = 0) -> List[str]:
        raw = str(text or "").strip()
        if not raw:
            return [""]

        words = raw.split()
        lines = []
        current = ""

        for word in words:
            candidate = (current + " " + word).strip()
            if pdf.stringWidth(candidate, font_name, font_size) <= max_width:
                current = candidate
            else:
                if current:
                    lines.append(current)
                current = word

        if current:
            lines.append(current)

        if max_lines and len(lines) > max_lines:
            lines = lines[:max_lines]
            if lines[-1] and len(lines[-1]) > 3:
                lines[-1] = lines[-1][:-3] + "..."

        return lines or [""]

    for idx, invoice in enumerate(invoices):
        voucher_x = 20 * mm
        voucher_y = 20 * mm
        voucher_w = page_width - (2 * voucher_x)
        voucher_h = page_height - (2 * voucher_y)

        # Slightly reduce inner padding to give the table a bit more breathing room.
        # This helps keep longer invoice numbers from looking cramped/overflowing.
        inner_pad = 6 * mm
        content_x = voucher_x + inner_pad
        content_w = voucher_w - (2 * inner_pad)
        top_y = voucher_y + voucher_h - inner_pad

        # Outer voucher border
        pdf.setLineWidth(0.8)
        pdf.rect(voucher_x, voucher_y, voucher_w, voucher_h)

        # Supplier box
        supplier_name = getattr(invoice, "supplier_name", None)
        if not supplier_name:
            supplier = getattr(invoice, "supplier", None)
            supplier_name = getattr(supplier, "name", None) if supplier else None
        supplier_name = supplier_name or "Unknown Supplier"

        sup_w = 62 * mm
        sup_h = 16 * mm
        sup_x = content_x
        sup_y = top_y - sup_h - (4 * mm)
        pdf.rect(sup_x, sup_y, sup_w, sup_h)

        pdf.setFont("Helvetica-Bold", 11)
        for i, line in enumerate(wrap_lines(supplier_name.upper(), sup_w - 4 * mm, "Helvetica-Bold", 11, max_lines=2)):
            line_y = sup_y + sup_h - 5 * mm - (i * 4.4 * mm)
            pdf.drawCentredString(sup_x + (sup_w / 2), line_y, line)

        # Voucher heading/date
        header_right_x = content_x + content_w
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawRightString(header_right_x, top_y - 1 * mm, "PAYMENT VOUCHER")
        pdf.setFont("Helvetica", 9)
        pdf.drawRightString(
            header_right_x,
            top_y - 7 * mm,
            f"Date: {format_date_value(getattr(invoice, 'invoice_date', None))}"
        )

        # Voucher table
        table_x = content_x
        table_y_top = sup_y - 7 * mm
        # Use the full available content width (158mm with current margins/padding).
        # Widen the "Invoice" column since some invoice numbers are long (e.g. 15-16 chars without spaces).
        col_widths = [18 * mm, 15 * mm, 30 * mm, 45 * mm, 25 * mm, 25 * mm]
        headers = ["PJV No", "TF No", "Invoice", "Description", "Amount/Invoice", "Amount Paid"]

        desc_text = getattr(invoice, "description", "") or ""
        desc_lines = wrap_lines(desc_text, col_widths[3] - 3 * mm, "Helvetica", 8.5, max_lines=5)
        body_h = max(10 * mm, (len(desc_lines) * 4.0 * mm) + 3 * mm)
        header_h = 8 * mm
        table_h = header_h + body_h

        # Draw table frame
        pdf.rect(table_x, table_y_top - table_h, sum(col_widths), table_h)
        pdf.line(table_x, table_y_top - header_h, table_x + sum(col_widths), table_y_top - header_h)

        # Vertical separators
        cursor_x = table_x
        for width in col_widths[:-1]:
            cursor_x += width
            pdf.line(cursor_x, table_y_top, cursor_x, table_y_top - table_h)

        # Header labels
        pdf.setFont("Helvetica-Bold", 8)
        cursor_x = table_x
        for i, label in enumerate(headers):
            center_x = cursor_x + (col_widths[i] / 2)
            pdf.drawCentredString(center_x, table_y_top - 5.5 * mm, label)
            cursor_x += col_widths[i]

        tf_or_chq = getattr(invoice, "tf_number", None) or getattr(invoice, "chq_number", None) or "-"
        invoice_values = [
            str(getattr(invoice, "pjv_number", "") or ""),
            str(tf_or_chq),
            str(getattr(invoice, "invoice_number", "") or ""),
            desc_text,
            f"EUR {float(getattr(invoice, 'invoice_amount', 0) or 0):,.2f}",
            f"EUR {float(getattr(invoice, 'payment_amount', 0) or 0):,.2f}",
        ]

        # Body values
        cursor_x = table_x
        value_top_y = table_y_top - header_h - 3.2 * mm
        pdf.setFont("Helvetica", 8.5)
        for i, value in enumerate(invoice_values):
            if i == 3:
                for line_idx, line in enumerate(desc_lines):
                    pdf.drawString(cursor_x + 1.5 * mm, value_top_y - (line_idx * 4.0 * mm), line)
            elif i in (4, 5):
                pdf.drawRightString(cursor_x + col_widths[i] - 1.5 * mm, value_top_y, value)
            else:
                lines = wrap_lines(value, col_widths[i] - 3 * mm, "Helvetica", 8.5, max_lines=2)
                for line_idx, line in enumerate(lines):
                    # Center these short identifier fields so the table feels balanced.
                    center_x = cursor_x + (col_widths[i] / 2)
                    pdf.drawCentredString(center_x, value_top_y - (line_idx * 4.0 * mm), line)
            cursor_x += col_widths[i]

        # Footer text and signature area
        footer_y = voucher_y + 30 * mm
        pdf.setFont("Helvetica-Bold", 8.5)
        pdf.drawString(content_x, footer_y + 10 * mm, "Kindly remit fiscal receipt and statement at the end of the month")

        pdf.line(content_x, footer_y - 2 * mm, content_x + 48 * mm, footer_y - 2 * mm)
        pdf.setFont("Helvetica", 8.5)
        pdf.drawString(content_x, footer_y - 7 * mm, "Executive Secretary")

        if idx < len(invoices) - 1:
            pdf.showPage()

    pdf.save()
    buffer.seek(0)
    return buffer


def get_export_filename(file_type: str, approved_only: bool = False, prefix: str = "Schedule_of_Payments") -> str:
    """Generate filename for export."""
    now = datetime.now()
    month_year = now.strftime("%B_%Y")
    suffix = "_Approved" if approved_only else "_All"
    return f"{prefix}_{month_year}{suffix}.{file_type}"
