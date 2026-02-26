"""
Pytest tests for export functionality with date range filtering.

Tests PDF/Excel/CSV exports with date_from and date_to parameters.
"""
import pytest
import re
import asyncio
from types import SimpleNamespace
from datetime import datetime, date
from openpyxl import load_workbook
from database import get_db


@pytest.mark.unit
def test_build_period_label_with_dates():
    """Test period label building with date range."""
    from routes.exports import build_period_label

    label = build_period_label("2026-01-01", "2026-01-31")
    assert label == "2026-01-01 to 2026-01-31"


@pytest.mark.unit
def test_build_period_label_with_only_date_from():
    """Test period label building with only start date."""
    from routes.exports import build_period_label

    label = build_period_label("2026-01-01", None)
    assert label == "2026-01-01 to ..."


@pytest.mark.unit
def test_build_period_label_with_only_date_to():
    """Test period label building with only end date."""
    from routes.exports import build_period_label

    label = build_period_label(None, "2026-01-31")
    assert label == "... to 2026-01-31"


@pytest.mark.unit
def test_build_period_label_without_dates():
    """Test period label defaults to current month when no dates provided."""
    from routes.exports import build_period_label

    label = build_period_label(None, None)
    # Should be current month/year like "February 2026"
    current_month_year = datetime.now().strftime("%B %Y")
    assert label == current_month_year


@pytest.mark.unit
def test_build_signatory_overrides():
    """Test signatory override helper output format."""
    from routes.exports import build_signatory_overrides

    overrides = build_signatory_overrides(
        sindku="SINDKU NAME",
        segretarju_ezekuttiv="SECRETARY NAME",
        proponent="PROPONENT NAME",
        sekondant="SEKONDANT NAME",
    )

    assert overrides == {
        "sindku": "SINDKU NAME",
        "segretarju_ezekuttiv": "SECRETARY NAME",
        "proponent": "PROPONENT NAME",
        "sekondant": "SEKONDANT NAME",
    }


@pytest.mark.integration
def test_export_signatory_profile_persistence():
    """Signatory values should be normalized and persisted in settings."""
    from services.export_profile_service import (
        SIGNATORY_SETTINGS_KEYS,
        resolve_export_signatories,
        get_export_signatories,
    )

    keys = tuple(SIGNATORY_SETTINGS_KEYS.values())
    placeholders = ",".join("?" for _ in keys)

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(f"DELETE FROM settings WHERE key IN ({placeholders})", keys)
        conn.commit()

        try:
            updated = resolve_export_signatories(conn, {
                "sindku": "  Maria   Azzopardi  ",
                "segretarju_ezekuttiv": " Joseph Camilleri ",
                "proponent": " Anne   Borg ",
                "sekondant": " Luke   Vella ",
            })

            assert updated["sindku"] == "Maria Azzopardi"
            assert updated["segretarju_ezekuttiv"] == "Joseph Camilleri"
            assert updated["proponent"] == "Anne Borg"
            assert updated["sekondant"] == "Luke Vella"

            loaded = get_export_signatories(conn)
            assert loaded == updated
        finally:
            cursor.execute(f"DELETE FROM settings WHERE key IN ({placeholders})", keys)
            conn.commit()


@pytest.mark.unit
def test_generate_schedule_excel_populates_signatory_names():
    """Excel export should include configured signatory names in signature section."""
    from services.export_service import generate_schedule_excel

    invoice = SimpleNamespace(
        supplier_id=1,
        supplier=SimpleNamespace(name="Supplier One"),
        invoice_amount=100.0,
        payment_amount=100.0,
        method_request="Inv",
        method_procurement="DA",
        description="Test invoice",
        invoice_date=date(2026, 2, 1),
        invoice_number="INV-001",
        po_number="PO-001",
        chq_number=None,
        tf_number="10/2026",
        pjv_number="PJV-001",
        is_deleted=0,
        is_void=0,
    )

    signatories = {
        "sindku": "Sindku Person",
        "segretarju_ezekuttiv": "Secretary Person",
        "proponent": "Proponent Person",
        "sekondant": "Sekondant Person",
    }

    buffer = generate_schedule_excel(
        invoices=[invoice],
        sitting_number="12",
        month_year="2026-02-01 to 2026-02-28",
        signatories=signatories,
    )

    wb = load_workbook(buffer)
    ws = wb.active
    all_values = {
        str(cell.value)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=16)
        for cell in row
        if cell.value is not None
    }

    assert "Sindku Person" in all_values
    assert "Secretary Person" in all_values
    assert "Proponent Person" in all_values
    assert "Sekondant Person" in all_values


@pytest.mark.integration
def test_get_invoices_for_export_with_date_range():
    """Test that get_invoices_for_export filters by date range correctly."""
    from routes.exports import get_invoices_for_export

    with get_db() as conn:
        cursor = conn.cursor()

        # Create test data with different dates
        cursor.execute("DELETE FROM invoices WHERE pjv_number LIKE 'TEST-DATE-%'")

        # Invoice within range
        cursor.execute("""
            INSERT INTO invoices (
                supplier_id, invoice_amount, payment_amount, method_request,
                method_procurement, description, invoice_date, invoice_number,
                pjv_number, is_approved, is_deleted
            ) VALUES (1, 100, 100, 'Inv', 'DA', 'Test 1', '2026-01-15', 'INV-1', 'TEST-DATE-1', 0, 0)
        """)

        # Invoice before range
        cursor.execute("""
            INSERT INTO invoices (
                supplier_id, invoice_amount, payment_amount, method_request,
                method_procurement, description, invoice_date, invoice_number,
                pjv_number, is_approved, is_deleted
            ) VALUES (1, 200, 200, 'Inv', 'DA', 'Test 2', '2025-12-01', 'INV-2', 'TEST-DATE-2', 0, 0)
        """)

        # Invoice after range
        cursor.execute("""
            INSERT INTO invoices (
                supplier_id, invoice_amount, payment_amount, method_request,
                method_procurement, description, invoice_date, invoice_number,
                pjv_number, is_approved, is_deleted
            ) VALUES (1, 300, 300, 'Inv', 'DA', 'Test 3', '2026-02-15', 'INV-3', 'TEST-DATE-3', 0, 0)
        """)

        # Test date range filter
        invoices = get_invoices_for_export(
            conn,
            date_from="2026-01-01",
            date_to="2026-01-31"
        )

        # Should only get the invoice from 2026-01-15
        test_invoices = [inv for inv in invoices if inv.pjv_number.startswith('TEST-DATE-')]
        assert len(test_invoices) == 1
        assert test_invoices[0].pjv_number == 'TEST-DATE-1'

        # Cleanup
        cursor.execute("DELETE FROM invoices WHERE pjv_number LIKE 'TEST-DATE-%'")


@pytest.mark.integration
def test_get_invoices_for_export_date_from_only():
    """Test export filtering with only date_from."""
    from routes.exports import get_invoices_for_export

    with get_db() as conn:
        cursor = conn.cursor()

        cursor.execute("DELETE FROM invoices WHERE pjv_number LIKE 'TEST-DATEFROM-%'")

        # Old invoice
        cursor.execute("""
            INSERT INTO invoices (
                supplier_id, invoice_amount, payment_amount, method_request,
                method_procurement, description, invoice_date, invoice_number,
                pjv_number, is_approved, is_deleted
            ) VALUES (1, 100, 100, 'Inv', 'DA', 'Old', '2025-12-01', 'INV-OLD', 'TEST-DATEFROM-1', 0, 0)
        """)

        # New invoice
        cursor.execute("""
            INSERT INTO invoices (
                supplier_id, invoice_amount, payment_amount, method_request,
                method_procurement, description, invoice_date, invoice_number,
                pjv_number, is_approved, is_deleted
            ) VALUES (1, 200, 200, 'Inv', 'DA', 'New', '2026-02-01', 'INV-NEW', 'TEST-DATEFROM-2', 0, 0)
        """)

        # Filter from 2026-01-01 onwards
        invoices = get_invoices_for_export(conn, date_from="2026-01-01")

        # Should only get new invoice
        test_invoices = [inv for inv in invoices if inv.pjv_number.startswith('TEST-DATEFROM-')]
        assert len(test_invoices) == 1
        assert test_invoices[0].pjv_number == 'TEST-DATEFROM-2'

        # Cleanup
        cursor.execute("DELETE FROM invoices WHERE pjv_number LIKE 'TEST-DATEFROM-%'")


@pytest.mark.integration
def test_export_filters_exclude_void_by_default():
    """Test that exports exclude voided invoices by default."""
    from routes.exports import get_invoices_for_export

    with get_db() as conn:
        cursor = conn.cursor()

        cursor.execute("DELETE FROM invoices WHERE pjv_number LIKE 'TEST-VOID-%'")

        # Normal invoice
        cursor.execute("""
            INSERT INTO invoices (
                supplier_id, invoice_amount, payment_amount, method_request,
                method_procurement, description, invoice_date, invoice_number,
                pjv_number, is_approved, is_deleted, is_void
            ) VALUES (1, 100, 100, 'Inv', 'DA', 'Normal', '2026-01-15', 'INV-N', 'TEST-VOID-1', 0, 0, 0)
        """)

        # Voided invoice
        cursor.execute("""
            INSERT INTO invoices (
                supplier_id, invoice_amount, payment_amount, method_request,
                method_procurement, description, invoice_date, invoice_number,
                pjv_number, is_approved, is_deleted, is_void, void_reason
            ) VALUES (1, 200, 200, 'Inv', 'DA', 'Voided', '2026-01-16', 'INV-V', 'TEST-VOID-2', 0, 0, 1, 'Test void')
        """)

        # Default: exclude void
        invoices = get_invoices_for_export(conn, include_void=False)
        test_invoices = [inv for inv in invoices if inv.pjv_number.startswith('TEST-VOID-')]
        assert len(test_invoices) == 1
        assert test_invoices[0].pjv_number == 'TEST-VOID-1'

        # Include void explicitly
        invoices_with_void = get_invoices_for_export(conn, include_void=True)
        test_invoices_with_void = [inv for inv in invoices_with_void if inv.pjv_number.startswith('TEST-VOID-')]
        assert len(test_invoices_with_void) == 2

        # Cleanup
        cursor.execute("DELETE FROM invoices WHERE pjv_number LIKE 'TEST-VOID-%'")


@pytest.mark.unit
def test_infer_period_from_invoices():
    """Test that period can be inferred from invoice dates."""
    from routes.exports import infer_period_from_invoices
    from types import SimpleNamespace

    # Create mock invoice objects
    invoices = [
        SimpleNamespace(invoice_date=date(2026, 1, 15)),
        SimpleNamespace(invoice_date=date(2026, 1, 25)),
        SimpleNamespace(invoice_date=date(2026, 1, 5))
    ]

    date_from, date_to = infer_period_from_invoices(invoices)

    assert date_from == "2026-01-05"  # Minimum date
    assert date_to == "2026-01-25"    # Maximum date


@pytest.mark.unit
def test_infer_period_handles_string_dates():
    """Test that period inference handles date strings."""
    from routes.exports import infer_period_from_invoices
    from types import SimpleNamespace

    # Create mock invoices with string dates
    invoices = [
        SimpleNamespace(invoice_date="2026-01-15"),
        SimpleNamespace(invoice_date="2026-01-25")
    ]

    date_from, date_to = infer_period_from_invoices(invoices)

    assert date_from == "2026-01-15"
    assert date_to == "2026-01-25"


@pytest.mark.unit
def test_infer_period_returns_none_for_empty():
    """Test that period inference returns None for empty invoice list."""
    from routes.exports import infer_period_from_invoices

    date_from, date_to = infer_period_from_invoices([])

    assert date_from is None
    assert date_to is None


@pytest.mark.unit
def test_generate_bulk_vouchers_pdf_has_one_page_per_invoice():
    """Bulk voucher export should generate one PDF page per invoice."""
    from services.export_service import generate_bulk_vouchers_pdf

    invoices = [
        SimpleNamespace(
            supplier_name="Supplier One",
            invoice_date=date(2026, 2, 1),
            pjv_number="PJV-TEST-1",
            tf_number="40/2026",
            chq_number=None,
            invoice_number="INV-TEST-1",
            description="First test invoice",
            invoice_amount=123.45,
            payment_amount=123.45
        ),
        SimpleNamespace(
            supplier_name="Supplier Two",
            invoice_date=date(2026, 2, 2),
            pjv_number="PJV-TEST-2",
            tf_number=None,
            chq_number="20/2026",
            invoice_number="INV-TEST-2",
            description="Second test invoice",
            invoice_amount=678.90,
            payment_amount=678.90
        )
    ]

    pdf_buffer = generate_bulk_vouchers_pdf(invoices)
    payload = pdf_buffer.getvalue()

    assert payload.startswith(b"%PDF"), "Expected a valid PDF payload"
    assert len(re.findall(rb"/Type\s*/Page\b", payload)) >= 2


@pytest.mark.integration
def test_export_selected_voucher_pdf_endpoint_returns_pdf():
    """Selected voucher export endpoint should return a multi-page PDF."""
    from routes.exports import export_selected_voucher_pdf

    class FakeRequest:
        def __init__(self, payload):
            self._payload = payload
            self.state = SimpleNamespace(user_id=1)
            self.client = SimpleNamespace(host="127.0.0.1")

        async def json(self):
            return self._payload

    async def collect_stream(response):
        chunks = []
        async for chunk in response.body_iterator:
            chunks.append(chunk)
        return b"".join(chunks)

    created_ids = []
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM invoices WHERE pjv_number LIKE 'TEST-VOUCHER-EXPORT-%'")

        for idx in range(2):
            cursor.execute(
                """
                INSERT INTO invoices (
                    supplier_id, invoice_amount, payment_amount, method_request,
                    method_procurement, description, invoice_date, invoice_number,
                    pjv_number, is_approved, is_deleted, is_void, tf_number
                ) VALUES (1, 100, 100, 'Inv', 'DA', ?, '2026-02-01', ?, ?, 1, 0, 0, ?)
                """,
                (
                    f"Voucher export test {idx}",
                    f"INV-VOUCHER-{idx}",
                    f"TEST-VOUCHER-EXPORT-{idx}",
                    f"{80 + idx}/2026",
                ),
            )
            created_ids.append(cursor.lastrowid)

    try:
        request = FakeRequest({"invoice_ids": created_ids})
        response = asyncio.run(export_selected_voucher_pdf(request))
        pdf_bytes = asyncio.run(collect_stream(response))

        assert response.status_code == 200
        assert response.media_type == "application/pdf"
        assert response.headers.get("content-disposition", "").endswith(".pdf")
        assert pdf_bytes.startswith(b"%PDF")
        assert len(re.findall(rb"/Type\s*/Page\b", pdf_bytes)) >= 2
    finally:
        if created_ids:
            with get_db() as conn:
                cursor = conn.cursor()
                placeholders = ",".join("?" for _ in created_ids)
                cursor.execute(f"DELETE FROM invoices WHERE id IN ({placeholders})", created_ids)


@pytest.mark.integration
def test_export_pdf_persists_signatory_overrides():
    """PDF export endpoint should persist provided signatory names."""
    from routes.exports import export_pdf
    from services.export_profile_service import SIGNATORY_SETTINGS_KEYS, get_export_signatories

    class FakeRequest:
        def __init__(self):
            self.state = SimpleNamespace(user_id=1)
            self.client = SimpleNamespace(host="127.0.0.1")

    async def collect_stream(response):
        chunks = []
        async for chunk in response.body_iterator:
            chunks.append(chunk)
        return b"".join(chunks)

    keys = tuple(SIGNATORY_SETTINGS_KEYS.values())
    placeholders = ",".join("?" for _ in keys)

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(f"DELETE FROM settings WHERE key IN ({placeholders})", keys)
        conn.commit()

    try:
        request = FakeRequest()
        response = asyncio.run(
            export_pdf(
                request=request,
                approved_only=False,
                sitting_number=None,
                status=None,
                supplier_id=None,
                date_from="2026-01-01",
                date_to="2026-01-31",
                include_void=False,
                sindku="Mayor Name",
                segretarju_ezekuttiv="Executive Secretary Name",
                proponent="Proponent Name",
                sekondant="Sekondant Name",
            )
        )
        pdf_bytes = asyncio.run(collect_stream(response))

        assert response.status_code == 200
        assert response.media_type == "application/pdf"
        assert pdf_bytes.startswith(b"%PDF")

        with get_db() as conn:
            stored = get_export_signatories(conn)
            assert stored["sindku"] == "Mayor Name"
            assert stored["segretarju_ezekuttiv"] == "Executive Secretary Name"
            assert stored["proponent"] == "Proponent Name"
            assert stored["sekondant"] == "Sekondant Name"
    finally:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute(f"DELETE FROM settings WHERE key IN ({placeholders})", keys)
            conn.commit()


if __name__ == "__main__":
    # Allow running this file directly
    pytest.main([__file__, "-v"])
